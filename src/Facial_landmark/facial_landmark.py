import cv2
import mediapipe as mp
import numpy as np
import os
from openpyxl import Workbook

mp_face_mesh = mp.solutions.face_mesh

def create_faciallandmark():
    return Faciallandmark()

class Faciallandmark:
    def __init__(self) -> None:
        pass

    def get_landmarks(self,imgpath):
        """
        @brief: Obtain facial landmarks values.
        @param:
            imagepath: The path of the image that needs to be processed.
        @return:
            coords: The obtained landmarks.
        """
        with mp_face_mesh.FaceMesh(static_image_mode=True,max_num_faces=1,refine_landmarks=True,min_detection_confidence=0.5) as face_mesh:
            image = cv2.imdecode(np.fromfile(imgpath, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
            results = face_mesh.process(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
            try:
                coords=np.array(results.multi_face_landmarks[0].landmark)
                return coords
            except:
                print('Face detection failed.')
                return []
            
            
    def get_landmarks_batching(self,directorypath):
        """
        @brief: Batch obtaining image landmarks and save as xlsx file.
        @param:
            directorypath: Path to the image folder.
        """
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'subject ID'
        ws['B1'] = 'X'
        # ws['C1'].value = 'score'
        ws['C1'] = 'Y'
        ws['D1'] = 'Z'
        for filename in os.listdir(directorypath):
            num = 1
            face_file_path = (directorypath + "/" + filename)
            with mp_face_mesh.FaceMesh(static_image_mode=True,max_num_faces=1,refine_landmarks=True,min_detection_confidence=0.5) as face_mesh:
                coords = self.get_landmarks(face_file_path)
            if len(coords) != 0:
                for i in coords:
                    num_ex = str(num)
                    x = str(i.x)
                    y = str(i.y)
                    z = str(i.z)
                    row = [num_ex,x,y,z]
                    ws.append(row)
                    num += 1
            else:
                continue
        wb.save('result.xlsx')